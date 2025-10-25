"""
Close Cash Excel Management Views
"""
import os
import json
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse, Http404
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.conf import settings
from django.utils import timezone
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import mimetypes
from django.views.decorators.http import require_POST
from django.db import transaction

from .models import CloseCashSchema, CloseCashEntry, A2ZSnapshot
from .close_cash_excel import (
    get_close_cash_directory,
    build_workbook_schema_and_data,
    write_values_to_workbook,
)


def is_employee(user):
    """Check if user is an employee (staff member)."""
    return user.is_staff


def get_close_cash_directory():
    """Get the Close Cash directory path."""
    return os.path.join(settings.BASE_DIR, 'Close Cash')


def validate_filename(filename):
    """Validate filename to prevent path traversal attacks."""
    if not filename or '..' in filename or '/' in filename or '\\' in filename:
        return False
    return filename.endswith('.xlsx') and len(filename) <= 50


def get_user_excel_files(user):
    """Get Excel files available to the user based on their role."""
    close_cash_dir = get_close_cash_directory()
    
    if not os.path.exists(close_cash_dir):
        return []
    
    files = []
    for filename in os.listdir(close_cash_dir):
        if filename.endswith('.xlsx') and validate_filename(filename):
            # Admin can see all files
            if user.is_superuser:
                files.append({
                    'filename': filename,
                    'display_name': filename.replace('.xlsx', ''),
                    'is_master': filename == 'A to Z Format.xlsx',
                    'path': os.path.join(close_cash_dir, filename)
                })
            # Employee can only see their own file
            elif filename.lower() == f"{user.username}.xlsx":
                files.append({
                    'filename': filename,
                    'display_name': filename.replace('.xlsx', ''),
                    'is_master': False,
                    'path': os.path.join(close_cash_dir, filename)
                })
    
    return files


@login_required
def close_cash_dashboard(request):
    """Main dashboard showing available Excel files based on user role."""
    if not is_employee(request.user):
        messages.error(request, 'You do not have employee access.')
        return redirect('index')

    # Show employee entrypoint for employees/admin, blank for others
    context = {}
    if is_employee_or_admin(request.user):
        context['show_employee_entry'] = True
    
    return render(request, 'accounts/close_cash_dashboard.html', context)


@login_required
def view_excel_file(request, filename):
    """Display Excel file as editable HTML table."""
    if not is_employee(request.user):
        messages.error(request, 'You do not have employee access.')
        return redirect('index')

    # Restrict legacy grid editor entirely by redirecting to blank page
    messages.info(request, 'Close Cash editor is currently unavailable.')
    return redirect('close_cash_dashboard')
    
    if not validate_filename(filename):
        messages.error(request, 'Invalid filename.')
        return redirect('close_cash_dashboard')
    
    # Check if user has access to this file
    user_files = get_user_excel_files(request.user)
    file_info = None
    for f in user_files:
        if f['filename'] == filename:
            file_info = f
            break
    
    if not file_info:
        messages.error(request, 'You do not have access to this file.')
        return redirect('close_cash_dashboard')
    
    file_path = file_info['path']
    
    if not os.path.exists(file_path):
        messages.error(request, 'File not found.')
        return redirect('close_cash_dashboard')
    
    try:
        # Load workbook with data_only=True to get calculated values
        workbook_data = load_workbook(file_path, data_only=True)  # For values
        workbook_formulas = load_workbook(file_path, data_only=False)  # For formulas
        
        # Get all sheet names
        sheet_names = workbook_data.sheetnames
        
        # Load data for all sheets
        all_sheets_data = {}
        for sheet_name in sheet_names:
            ws_data = workbook_data[sheet_name]
            ws_formulas = workbook_formulas[sheet_name]
            
            sheet_data = []
            max_row = ws_data.max_row or 1
            max_col = ws_data.max_column or 1
            
            for row in range(1, max_row + 1):
                row_data = []
                for col in range(1, max_col + 1):
                    cell_data = ws_data.cell(row=row, column=col)
                    cell_formula = ws_formulas.cell(row=row, column=col)
                    
                    # Show calculated value, but track if it's a formula
                    value = cell_data.value if cell_data.value is not None else ''
                    has_formula = str(cell_formula.value).startswith('=') if cell_formula.value else False
                    
                    row_data.append({
                        'value': value,
                        'hasFormula': has_formula
                    })
                sheet_data.append(row_data)
            
            # Ensure data has valid structure - add empty row if sheet is completely empty
            if not sheet_data or len(sheet_data) == 0:
                # Add at least one empty row if sheet is completely empty
                sheet_data = [[{'value': '', 'hasFormula': False}] * 10]
            
            all_sheets_data[sheet_name] = sheet_data
        
        # Get file stats
        file_stats = os.stat(file_path)
        last_modified = timezone.datetime.fromtimestamp(file_stats.st_mtime)
        
        context = {
            'filename': filename,
            'display_name': file_info['display_name'],
            'is_master': file_info['is_master'],
            'all_sheets_data': json.dumps(all_sheets_data),
            'sheet_names': json.dumps(sheet_names),  # ADD JSON SERIALIZATION
            'active_sheet': sheet_names[0],
            'last_modified': last_modified,
            'file_size': file_stats.st_size,
        }
        
        return render(request, 'accounts/close_cash_editor.html', context)
        
    except Exception as e:
        messages.error(request, f'Error reading Excel file: {str(e)}')
        return redirect('close_cash_dashboard')


# =====================
# DB-backed Forms (New)
# =====================

@login_required
def close_cash_form_dashboard(request):
    if not is_employee(request.user):
        messages.error(request, 'You do not have employee access.')
        return redirect('index')

    # Determine employee workbook name by username.xlsx
    employee_filename = f"{request.user.username}.xlsx".lower()
    schemas = CloseCashSchema.objects.filter(workbook__iexact=employee_filename).order_by('sheet_name')
    entries = CloseCashEntry.objects.filter(user=request.user, workbook__iexact=employee_filename)
    latest_by_sheet = {e.sheet_name: e for e in entries}

    context = {
        'schemas': schemas,
        'latest_by_sheet': latest_by_sheet,
        'employee_filename': employee_filename,
    }
    return render(request, 'accounts/close_cash_form_dashboard.html', context)


@login_required
def edit_close_cash_form(request, sheet_name):
    if not is_employee(request.user):
        messages.error(request, 'You do not have employee access.')
        return redirect('index')

    employee_filename = f"{request.user.username}.xlsx".lower()
    try:
        schema_obj = CloseCashSchema.objects.get(workbook__iexact=employee_filename, sheet_name=sheet_name, version='v1')
    except CloseCashSchema.DoesNotExist:
        messages.error(request, 'Form schema not found for this sheet.')
        return redirect('close_cash_forms_dashboard')

    # Load latest entry for prefill
    entry = CloseCashEntry.objects.filter(
        user=request.user,
        workbook__iexact=employee_filename,
        sheet_name=sheet_name,
        source_version='v1'
    ).order_by('-created_at').first()

    context = {
        'sheet_name': sheet_name,
        'schema': schema_obj.schema_json,
        'values': entry.data_json if entry else {},
    }
    return render(request, 'accounts/close_cash_form_edit.html', context)


@login_required
@require_POST
@transaction.atomic
def submit_close_cash_form(request, sheet_name):
    if not is_employee(request.user):
        return JsonResponse({'error': 'Access denied'}, status=403)

    try:
        payload = json.loads(request.body.decode('utf-8'))
    except Exception:
        return JsonResponse({'error': 'Invalid JSON payload'}, status=400)

    data = payload.get('data', {})
    entry_date = payload.get('entry_date')  # ISO date string optional
    employee_filename = f"{request.user.username}.xlsx".lower()

    try:
        schema_obj = CloseCashSchema.objects.get(workbook__iexact=employee_filename, sheet_name=sheet_name, version='v1')
    except CloseCashSchema.DoesNotExist:
        return JsonResponse({'error': 'Form schema not found'}, status=404)

    # Convert entry_date
    from django.utils import timezone
    if entry_date:
        try:
            from datetime import date
            entry_date_obj = timezone.datetime.fromisoformat(entry_date).date()
        except Exception:
            return JsonResponse({'error': 'Invalid entry_date'}, status=400)
    else:
        entry_date_obj = timezone.now().date()

    # Persist DB entry
    entry = CloseCashEntry.objects.create(
        user=request.user,
        workbook=employee_filename,
        sheet_name=sheet_name,
        entry_date=entry_date_obj,
        data_json=data,
        source_version='v1',
    )

    # Sync back to employee workbook
    try:
        employee_workbook_path = os.path.join(get_close_cash_directory(), employee_filename)
        write_values_to_workbook(employee_workbook_path, sheet_name, schema_obj.schema_json, data)
    except Exception as e:
        # Do not rollback DB write; report sync failure
        return JsonResponse({'success': True, 'warning': f'Data saved but Excel sync failed: {str(e)}'})

    return JsonResponse({'success': True, 'message': 'Form saved successfully'})


# =====================
# A to Z Master Editor
# =====================

@login_required
def a2z_master_editor(request):
    if not is_employee(request.user):
        messages.error(request, 'You do not have employee access.')
        return redirect('index')

    master_filename = 'A to Z Format.xlsx'
    master_path = os.path.join(get_close_cash_directory(), master_filename)
    if not os.path.exists(master_path):
        if request.method == 'POST':
            return JsonResponse({'error': 'Master A to Z workbook not found.'}, status=404)
        else:
            messages.error(request, 'Master A to Z workbook not found.')
            return redirect('close_cash_dashboard')

    if request.method == 'POST':
        # Save edits back to master workbook
        try:
            payload = json.loads(request.body.decode('utf-8'))
            sheet_name = payload.get('sheet_name')
            data = payload.get('data') or {}
            if not sheet_name:
                return JsonResponse({'error': 'sheet_name is required'}, status=400)

            # Rebuild schema for the target sheet, then write values
            mapping = build_workbook_schema_and_data(master_path)
            sheet_payload = mapping.get(sheet_name)
            if not sheet_payload or 'schema' not in sheet_payload:
                return JsonResponse({'error': 'Schema not found for sheet'}, status=404)

            schema = sheet_payload['schema']
            write_values_to_workbook(master_path, sheet_name, schema, data)

            # Optional: snapshot full workbook after save
            try:
                updated_mapping = build_workbook_schema_and_data(master_path)
                A2ZSnapshot.objects.create(data_json=updated_mapping, note=f"Edited by {request.user.username}")
            except Exception:
                # Snapshot failure should not block save
                pass

            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'error': f'Failed to save: {str(e)}'}, status=500)

    # GET: render editor
    try:
        mapping = build_workbook_schema_and_data(master_path)
        return render(request, 'accounts/close_cash_a2z_editor.html', {
            'master_filename': master_filename,
            'mapping_json': json.dumps(mapping),
        })
    except Exception as e:
        messages.error(request, f'Error reading A to Z workbook: {str(e)}')
        return redirect('close_cash_dashboard')


# =====================
# Rawad-only Forms
# =====================

def is_employee_or_admin(user):
    """Check if user is a staff member or admin."""
    return user.is_staff or user.is_superuser


@login_required
def employee_forms_dashboard(request):
    """Employee forms dashboard showing available submissions."""
    if not is_employee_or_admin(request.user):
        messages.error(request, 'You do not have access to this page.')
        return redirect('index')
    
    # Get existing entries from DB - admin sees all, employee sees only their own
    if request.user.is_superuser:
        # Admin can see all submissions grouped by username
        entries = CloseCashEntry.objects.filter(
            workbook__iexact='Employee_Close_Cash.xlsx'
        ).order_by('user__username', '-entry_date', '-created_at')
        
        # Admin view: group by username, then by date
        from collections import defaultdict
        submissions_by_user = defaultdict(lambda: defaultdict(list))
        for entry in entries:
            username = entry.user.username
            date_key = entry.entry_date.strftime('%Y-%m-%d')
            submissions_by_user[username][date_key].append(entry)
        
        # Convert nested defaultdict to regular dict (both levels)
        submissions_by_user_dict = {
            username: dict(dates_dict) 
            for username, dates_dict in submissions_by_user.items()
        }
        
        context = {
            'submissions_by_user': submissions_by_user_dict,
            'is_admin': True,
            'workbook_name': 'Employee_Close_Cash.xlsx',
        }
    else:
        # Employee sees only their own submissions
        entries = CloseCashEntry.objects.filter(
            user=request.user, 
            workbook__iexact='Employee_Close_Cash.xlsx'
        ).order_by('-entry_date', '-created_at')
        
        # Employee view: group by date only
        from collections import defaultdict
        submissions_by_date = defaultdict(list)
        for entry in entries:
            date_key = entry.entry_date.strftime('%Y-%m-%d')
            submissions_by_date[date_key].append(entry)
        
        context = {
            'submissions_by_date': dict(submissions_by_date),
            'is_admin': False,
            'workbook_name': 'Employee_Close_Cash.xlsx',
        }
    return render(request, 'accounts/close_cash_employee_dashboard.html', context)


@login_required
def employee_edit_close_cash_form(request, sheet_name):
    """Edit employee form for specific date sheet."""
    if not is_employee_or_admin(request.user):
        messages.error(request, 'You do not have access to this page.')
        return redirect('index')
    
    # If sheet_name is 'export', redirect to export view
    # This shouldn't happen due to URL ordering, but as a safeguard
    if sheet_name == 'export':
        return redirect('employee_export_excel')
    
    from .close_cash_forms import (
        GeneralSectionForm, LebaneseCashForm, DollarCashForm, 
        SpecialCreditForm, CreditEntryFormSet, CoffeeMachineEntryFormSet
    )
    
    # Handle "new" sheet name for creating new submissions
    if sheet_name == 'new':
        sheet_name = 'new_submission'
        form_data = {}
    else:
        # Load existing entry from DB for prefill (if exists)
        entry = CloseCashEntry.objects.filter(
            workbook__iexact='Employee_Close_Cash.xlsx',
            sheet_name=sheet_name,
            source_version='v1'
        ).order_by('-created_at').first()
        
        # Use DB values if available, otherwise empty
        form_data = entry.data_json if entry else {}
    
    # Handle both old flat structure and new nested structure
    # If data has 'general' key, it's new structure; otherwise it's old flat structure
    if 'general' in form_data:
        # New nested structure
        general_initial = form_data.get('general', {})
        lebanese_cash_initial = form_data.get('lebanese_cash', {})
        dollar_cash_initial = form_data.get('dollar_cash', {})
        special_credit_initial = form_data.get('special_credit', {})
        credit_initial = form_data.get('credit', [])
        coffee_machine_initial = form_data.get('coffee_machine', [])
    else:
        # Old flat structure - convert to nested
        general_initial = {k: v for k, v in form_data.items() if k in ['black_market_daily_rate', 'cashier_name', 'date', 'shift_time']}
        lebanese_cash_initial = {k: v for k, v in form_data.items() if k.startswith('lebanese_')}
        dollar_cash_initial = {k: v for k, v in form_data.items() if k.startswith('dollar_')}
        special_credit_initial = {k: v for k, v in form_data.items() if k in ['rayan_invoices_credit', 'employee_invoice_credit', 'delivery_shabeb_co', 'delivery_employee', 'waste_goods']}
        credit_initial = form_data.get('credit', [])
        coffee_machine_initial = form_data.get('coffee_machine', [])
    
    # Initialize forms with data
    general_form = GeneralSectionForm(initial=general_initial)
    lebanese_cash_form = LebaneseCashForm(initial=lebanese_cash_initial)
    dollar_cash_form = DollarCashForm(initial=dollar_cash_initial)
    special_credit_form = SpecialCreditForm(initial=special_credit_initial)
    
    # Initialize formsets
    credit_formset = CreditEntryFormSet(initial=credit_initial, prefix='credit')
    coffee_machine_formset = CoffeeMachineEntryFormSet(initial=coffee_machine_initial, prefix='coffee')
    
    context = {
        'sheet_name': sheet_name,
        'general_form': general_form,
        'lebanese_cash_form': lebanese_cash_form,
        'dollar_cash_form': dollar_cash_form,
        'special_credit_form': special_credit_form,
        'credit_formset': credit_formset,
        'coffee_machine_formset': coffee_machine_formset,
        'workbook_name': 'Employee_Close_Cash.xlsx',
    }
    return render(request, 'accounts/close_cash_employee_form.html', context)


@login_required
@require_POST
@transaction.atomic
def employee_submit_close_cash_form(request, sheet_name):
    """Submit employee form data."""
    if not is_employee_or_admin(request.user):
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except Exception as e:
        return JsonResponse({'error': f'Invalid JSON payload: {str(e)}'}, status=400)
    
    data = payload.get('data', {})
    entry_date = payload.get('entry_date')
    
    # Convert entry_date
    from django.utils import timezone
    if entry_date:
        try:
            from datetime import date
            entry_date_obj = timezone.datetime.fromisoformat(entry_date).date()
        except Exception as e:
            return JsonResponse({'error': f'Invalid entry_date: {str(e)}'}, status=400)
    else:
        entry_date_obj = timezone.now().date()
    
    # Clean up dynamic list data - remove empty entries
    dynamic_list_keys = ['credit', 'coffee_machine']

    for key in dynamic_list_keys:
        if key in data and isinstance(data[key], list):
            if key == 'credit':
                # Filter out empty credit entries
                data[key] = [entry for entry in data[key] if entry and
                            (entry.get('amount') or entry.get('name'))]
            elif key == 'coffee_machine':
                # Filter out empty coffee machine entries
                data[key] = [entry for entry in data[key] if entry and
                            (entry.get('current_amount') or entry.get('daily_add') or entry.get('tag'))]

    # Validate required calculations are present
    calculated_fields = [
        'special_credit_total', 'lebanese_cash_total', 'dollar_cash_total_usd', 
        'dollar_cash_total_lbp', 'cash_purchase_total', 'credit_invoices_total',
        'employee_oth_total', 'customer_oth_total', 'bar_oth_total',
        'store_total', 'credit_grand_total', 'cash_in_hand_dollar', 'cash_in_hand_lebanese',
        'cash_out_of_hand', 'grand_total'
    ]
    
    for field in calculated_fields:
        if field not in data:
            data[field] = 0
    
    # Handle new vs existing submissions
    try:
        if sheet_name == 'new_submission' or sheet_name == 'new':
            # New submission - generate unique ID
            import uuid
            sheet_name = f"{entry_date_obj.strftime('%Y-%m-%d')}_{uuid.uuid4().hex[:8]}"
            
            entry = CloseCashEntry.objects.create(
                user=request.user,
                workbook='Employee_Close_Cash.xlsx',
                entry_date=entry_date_obj,
                sheet_name=sheet_name,
                data_json=data,
                source_version='v1',
            )
            created = True
        else:
            # Edit existing - find by sheet_name (allows admin to edit any submission)
            try:
                entry = CloseCashEntry.objects.get(
                    workbook='Employee_Close_Cash.xlsx',
                    sheet_name=sheet_name
                )
                entry.entry_date = entry_date_obj
                entry.data_json = data
                entry.save()
                created = False
            except CloseCashEntry.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': f'Submission with sheet_name "{sheet_name}" not found'
                }, status=404)
        
        action = 'created' if created else 'updated'
        
        # Return sheet_name and redirect URL for new submissions
        if created:
            from django.urls import reverse
            redirect_url = reverse('employee_edit_close_cash_form', args=[sheet_name])
            return JsonResponse({
                'success': True, 
                'message': f'Form {action} successfully',
                'sheet_name': sheet_name,
                'redirect_url': redirect_url
            })
        else:
            return JsonResponse({'success': True, 'message': f'Form {action} successfully'})
            
    except CloseCashEntry.DoesNotExist:
        return JsonResponse({'error': 'Submission not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Database error: {str(e)}'}, status=500)


@login_required
def employee_export_excel(request):
    """Export employee submissions as Excel file matching General_close_cash.xlsx format exactly."""
    if not is_employee_or_admin(request.user):
        messages.error(request, 'You do not have access to this page.')
        return redirect('index')
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        from django.contrib.auth.models import User
        from django.utils import timezone
        from io import BytesIO
        import logging
        
        logger = logging.getLogger(__name__)
        logger.info(f"Export request from user: {request.user.username}")
        
        # Determine which user's data to export
        username = request.GET.get('username', None)
        
        if username and request.user.is_superuser:
            # Admin exporting specific employee's data
            try:
                export_user = User.objects.get(username=username)
                logger.info(f"Admin exporting data for: {export_user.username}")
            except User.DoesNotExist:
                messages.error(request, f'User {username} not found.')
                return redirect('employee_forms_dashboard')
        else:
            # Export logged-in user's own data
            export_user = request.user
            logger.info(f"User exporting their own data: {export_user.username}")
        
        # Get submissions for the specific user
        entries = CloseCashEntry.objects.filter(
            user=export_user,
            workbook__iexact='Employee_Close_Cash.xlsx'
        ).order_by('entry_date')
        
        logger.info(f"Found {entries.count()} entries for user {export_user.username}")
        
        if not entries.exists():
            messages.error(request, f'No submissions found for {export_user.username}.')
            return redirect('employee_forms_dashboard')
        
        # Create new workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Process each submission
        for entry in entries:
            try:
                # Create sheet with date only
                sheet_name = entry.entry_date.strftime('%d-%m-%Y')
                ws = wb.create_sheet(title=sheet_name)
                
                data = entry.data_json
                logger.info(f"Processing entry for date: {sheet_name}")
                
                # Handle both old flat structure and new nested structure
                if 'general' in data:
                    # New nested structure
                    general_data = data.get('general', {})
                    lebanese_cash_data = data.get('lebanese_cash', {})
                    dollar_cash_data = data.get('dollar_cash', {})
                    special_credit_data = data.get('special_credit', {})
                else:
                    # Old flat structure
                    general_data = data
                    lebanese_cash_data = data
                    dollar_cash_data = data
                    special_credit_data = data
                
                # === SECTION 1: General Info + Lebanese/Dollar Cash (Rows 2-11) ===
                
                # General Info (A2-B11)
                ws['A2'] = "Cashier Name"
                ws['A3'] = general_data.get('cashier_name', '')
                ws['A4'] = "Date"
                ws['A5'] = general_data.get('date', '')
                ws['A7'] = "Cashier"
                ws['B7'] = general_data.get('cashier_name', '')
                ws['A8'] = "Date"
                ws['B8'] = general_data.get('date', '')
                ws['A9'] = "Shift"
                ws['B9'] = general_data.get('shift_time', '')
                
                # Lebanese Cash Headers (D2-H2)
                ws['D2'] = "Denomination"
                ws['E2'] = "#"
                ws['F2'] = "Value"
                ws['G2'] = "Dollar"
                ws['H2'] = "#"
                ws['I2'] = "Value"
                
                # Lebanese Cash Data (D3-G8)
                lbp_bills = [
                    (3, '5,000', lebanese_cash_data.get('lebanese_5000_qty', 0), 5000),
                    (4, '10,000', lebanese_cash_data.get('lebanese_10000_qty', 0), 10000),
                    (5, '20,000', lebanese_cash_data.get('lebanese_20000_qty', 0), 20000),
                    (6, '50,000', lebanese_cash_data.get('lebanese_50000_qty', 0), 50000),
                    (7, '100,000', lebanese_cash_data.get('lebanese_100000_qty', 0), 100000),
                ]
                
                for row_num, label, qty, denomination in lbp_bills:
                    ws[f'D{row_num}'] = label
                    ws[f'F{row_num}'] = qty
                    ws[f'G{row_num}'] = f"=F{row_num}*{denomination}"
                
                # Lebanese Total (D9-F9)
                ws['D9'] = "Total"
                ws['F9'] = f"=SUM(G3:G7)"
                
                # Dollar Cash Data (H3-J8)
                dollar_bills = [
                    (3, '1', dollar_cash_data.get('dollar_1_qty', 0), 1),
                    (4, '5', dollar_cash_data.get('dollar_5_qty', 0), 5),
                    (5, '10', dollar_cash_data.get('dollar_10_qty', 0), 10),
                    (6, '20', dollar_cash_data.get('dollar_20_qty', 0), 20),
                    (7, '50', dollar_cash_data.get('dollar_50_qty', 0), 50),
                    (8, '100', dollar_cash_data.get('dollar_100_qty', 0), 100),
                ]
                
                for row_num, label, qty, denomination in dollar_bills:
                    ws[f'H{row_num}'] = f"${label}"
                    ws[f'I{row_num}'] = qty
                    ws[f'J{row_num}'] = f"=I{row_num}*{denomination}"
                
                # Dollar Totals (H9-J10)
                ws['H9'] = "Total USD"
                ws['I9'] = f"=SUM(J3:J8)"
                ws['H10'] = "Total LBP"
                dollar_rate = dollar_cash_data.get('dollar_rate', 0)
                ws['I10'] = f"=I9*{dollar_rate}"
                
                # Special Credit (L4-M11)
                special_credit_items = [
                    (4, 'Rayan Invoices Credit', special_credit_data.get('rayan_invoices_credit', '')),
                    (5, 'Employee Invoice Credit', special_credit_data.get('employee_invoice_credit', '')),
                    (6, 'Delivery Shabeb co.', special_credit_data.get('delivery_shabeb_co', '')),
                    (7, 'Delivery Employee', special_credit_data.get('delivery_employee', '')),
                    (8, 'Waste Goods', special_credit_data.get('waste_goods', '')),
                ]
                
                for row_num, label, value in special_credit_items:
                    ws[f'L{row_num}'] = label
                    ws[f'M{row_num}'] = value
                
                # === SECTION 2: Credit Section (Rows 12-25) ===
                
                # Credit Headers (A12-I12)
                ws['A12'] = "Credit Entries"
                ws['D12'] = "Cash Purchase"
                ws['E12'] = "Credit Invoices"
                ws['F12'] = "Employee OTH"
                ws['G12'] = "Customer OTH"
                ws['H12'] = "Bar OTH"
                ws['I12'] = "Store"
                
                # Credit entries (A13-B24)
                credit_entries = data.get('credit', [])
                for i, entry_item in enumerate(credit_entries[:12]):  # Limit to 12 entries
                    row_num = 13 + i
                    ws[f'A{row_num}'] = entry_item.get('name', '')
                    ws[f'B{row_num}'] = f"{entry_item.get('amount', '')} {entry_item.get('currency', '')}"
                    
                    # Place in appropriate column based on tag
                    tag = entry_item.get('tag', '')
                    amount = entry_item.get('amount', 0)
                    if tag == 'Cash purchase':
                        ws[f'D{row_num}'] = amount
                    elif tag == 'Credit invoices':
                        ws[f'E{row_num}'] = amount
                    elif tag == 'Employee OTH':
                        ws[f'F{row_num}'] = amount
                    elif tag == 'Customer OTH':
                        ws[f'G{row_num}'] = amount
                    elif tag == 'Bar OTH':
                        ws[f'H{row_num}'] = amount
                    elif tag == 'Store':
                        ws[f'I{row_num}'] = amount
                
                # Credit Totals (A24-B24)
                ws['A24'] = "Credit Total"
                ws['B24'] = data.get('credit_grand_total', '')
                
                # Category Totals (C25-I25)
                ws['C25'] = "Totals"
                ws['D25'] = f"=SUM(D13:D24)"
                ws['E25'] = f"=SUM(E13:E24)"
                ws['F25'] = f"=SUM(F13:F24)"
                ws['G25'] = f"=SUM(G13:G24)"
                ws['H25'] = f"=SUM(H13:H24)"
                ws['I25'] = f"=SUM(I13:I24)"
                
                # === SECTION 3: Summary (Rows 26-31) ===
                
                # Cash in Hand (A26-B28)
                ws['A26'] = "Cash in Hand (Dollar)"
                ws['B26'] = data.get('cash_in_hand_dollar', '')
                ws['A27'] = "Cash in Hand (Lebanese)"
                ws['B27'] = data.get('cash_in_hand_lebanese', '')
                ws['A28'] = "Cash Out of Hand"
                ws['B28'] = data.get('cash_out_of_hand', '')
                
                # Grand Total (A29-B29)
                ws['A29'] = "Grand Total"
                ws['B29'] = data.get('grand_total', '')
                
                # Additional Special Credit (D27-I31)
                ws['D27'] = "Special Credit Total"
                ws['E27'] = data.get('special_credit_total', '')
                
                # === SECTION 4: Coffee Machine (Rows 36-72) ===
                
                # Coffee Machine Headers (A36-E36)
                ws['A36'] = "Coffee Machine Items"
                ws['E36'] = "Coffee Machine Items"
                
                # Coffee Machine entries (A40-E66)
                coffee_items = data.get('coffee_machine', [])
                for i, item in enumerate(coffee_items[:27]):  # Limit to 27 items
                    row_num = 40 + i
                    if item.get('tag'):  # Only show non-empty entries
                        ws[f'A{row_num}'] = item.get('tag', '')
                        ws[f'B{row_num}'] = f"Current: {item.get('current_amount', '')}"
                        ws[f'C{row_num}'] = f"Daily: {item.get('daily_add', '')}"
                        ws[f'E{row_num}'] = item.get('tag', '')
                        ws[f'F{row_num}'] = f"Current: {item.get('current_amount', '')}"
                        ws[f'G{row_num}'] = f"Daily: {item.get('daily_add', '')}"
                
                # Coffee Totals (A67-C67, E72-G72)
                ws['A67'] = "Coffee Total"
                ws['B67'] = "Total"
                ws['C67'] = "Total"
                ws['E72'] = "Coffee Total"
                ws['F72'] = "Total"
                ws['G72'] = "Total"
                
                # Apply formatting
                _apply_excel_formatting(ws)
                
            except Exception as sheet_error:
                logger.error(f"Error processing sheet {sheet_name}: {str(sheet_error)}")
                import traceback
                logger.error(f"Sheet traceback: {traceback.format_exc()}")
                continue
        
        # Convert to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return as download with username-based filename
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{export_user.username}_Close_Cash_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        logger.info(f"Successfully generated Excel file for {export_user.username}")
        return response
        
    except Exception as e:
        import traceback
        logger.error(f'Error generating Excel: {str(e)}')
        logger.error(f'Traceback: {traceback.format_exc()}')
        messages.error(request, f'Error generating Excel: {str(e)}')
        return redirect('employee_forms_dashboard')

def _apply_excel_formatting(ws):
    """Apply formatting to match General_close_cash.xlsx style."""
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    # Define styles
    header_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply headers formatting
    for row in range(2, 12):
        for col in range(1, 15):
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str) and any(keyword in str(cell.value).lower() for keyword in ['denomination', 'total', 'cashier', 'date', 'shift']):
                cell.font = header_font
    
    # Apply borders to data areas
    for row in range(2, 75):
        for col in range(1, 15):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                cell.border = border






@login_required
@require_http_methods(["POST"])
def save_excel_changes(request, filename):
    """Save edited data back to Excel file."""
    if not is_employee(request.user):
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    if not validate_filename(filename):
        return JsonResponse({'error': 'Invalid filename'}, status=400)
    
    # Check if user has access to this file
    user_files = get_user_excel_files(request.user)
    file_info = None
    for f in user_files:
        if f['filename'] == filename:
            file_info = f
            break
    
    if not file_info:
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        # Parse JSON data from request
        data = json.loads(request.body)
        changes = data.get('changes', [])
        
        if not changes:
            return JsonResponse({'success': True, 'message': 'No changes to save'})
        
        # Load Excel file with optimized settings
        file_path = file_info['path']
        workbook = load_workbook(file_path, data_only=False)  # Preserve formulas
        worksheet = workbook.active
        
        # Apply changes efficiently
        for change in changes:
            row = change.get('row')
            col = change.get('col')
            value = change.get('value')
            sheet_name = change.get('sheet', workbook.sheetnames[0])
            
            if row and col and value is not None:
                # Get the specific worksheet
                worksheet = workbook[sheet_name]
                cell = worksheet.cell(row=row, column=col)
                
                # Smart value conversion
                if value == '':
                    cell.value = None  # Empty cell
                elif isinstance(value, str):
                    # Check if it's a number
                    try:
                        if value.replace('.', '').replace('-', '').replace('+', '').isdigit():
                            cell.value = float(value)
                        else:
                            cell.value = value
                    except:
                        cell.value = value
                else:
                    cell.value = value
        
        # Save with optimized settings
        workbook.save(file_path)
        
        return JsonResponse({
            'success': True, 
            'message': f'Saved {len(changes)} changes to {filename}',
            'timestamp': timezone.now().isoformat()
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error saving file: {str(e)}'}, status=500)


@login_required
def download_excel_file(request, filename):
    """Download Excel file."""
    if not is_employee(request.user):
        messages.error(request, 'You do not have employee access.')
        return redirect('index')
    
    if not validate_filename(filename):
        messages.error(request, 'Invalid filename.')
        return redirect('close_cash_dashboard')
    
    # Check if user has access to this file
    user_files = get_user_excel_files(request.user)
    file_info = None
    for f in user_files:
        if f['filename'] == filename:
            file_info = f
            break
    
    if not file_info:
        messages.error(request, 'You do not have access to this file.')
        return redirect('close_cash_dashboard')
    
    file_path = file_info['path']
    
    if not os.path.exists(file_path):
        messages.error(request, 'File not found.')
        return redirect('close_cash_dashboard')
    
    try:
        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
    except Exception as e:
        messages.error(request, f'Error downloading file: {str(e)}')
        return redirect('close_cash_dashboard')


@login_required
@require_http_methods(["POST"])
def upload_excel_file(request, filename):
    """Upload modified Excel file."""
    if not is_employee(request.user):
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    if not validate_filename(filename):
        return JsonResponse({'error': 'Invalid filename'}, status=400)
    
    # Check if user has access to this file
    user_files = get_user_excel_files(request.user)
    file_info = None
    for f in user_files:
        if f['filename'] == filename:
            file_info = f
            break
    
    if not file_info:
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file uploaded'}, status=400)
        
        uploaded_file = request.FILES['file']
        
        # Validate file type
        if not uploaded_file.name.endswith('.xlsx'):
            return JsonResponse({'error': 'Only .xlsx files are allowed'}, status=400)
        
        # Save uploaded file
        file_path = file_info['path']
        with open(file_path, 'wb') as f:
            for chunk in uploaded_file.chunks():
                f.write(chunk)
        
        return JsonResponse({
            'success': True, 
            'message': f'Successfully uploaded {filename}',
            'timestamp': timezone.now().isoformat()
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error uploading file: {str(e)}'}, status=500)


@login_required
@require_POST
def employee_delete_submission(request, sheet_name):
    """Delete a specific submission (admin only)."""
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Access denied. Admin only.'}, status=403)
    
    try:
        # Find and delete the submission by sheet_name
        entry = CloseCashEntry.objects.get(
            workbook='Employee_Close_Cash.xlsx',
            sheet_name=sheet_name
        )
        
        username = entry.user.username
        entry_date = entry.entry_date.strftime('%Y-%m-%d')
        
        entry.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Submission deleted successfully for {username} on {entry_date}'
        })
        
    except CloseCashEntry.DoesNotExist:
        return JsonResponse({'error': 'Submission not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error deleting submission: {str(e)}'}, status=500)
