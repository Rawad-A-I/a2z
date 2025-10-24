import os
import re
from datetime import datetime
from typing import Dict, Any, List, Tuple, Optional

from django.conf import settings
from openpyxl import load_workbook


MASTER_FILENAME = "A to Z Format.xlsx"


def get_close_cash_directory() -> str:
    return os.path.join(settings.BASE_DIR, 'Close Cash')


def list_employee_workbooks() -> List[str]:
    directory = get_close_cash_directory()
    if not os.path.exists(directory):
        return []
    files = []
    for name in os.listdir(directory):
        if name.lower().endswith('.xlsx') and name != MASTER_FILENAME:
            files.append(os.path.join(directory, name))
    return files


def try_parse_date(value: str) -> Optional[datetime.date]:
    if not value:
        return None
    candidates = [
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%m-%d-%Y",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%Y/%m/%d",
        "%d %b %Y",
        "%d %B %Y",
        "%b %d, %Y",
        "%B %d, %Y",
    ]
    for fmt in candidates:
        try:
            return datetime.strptime(value.strip(), fmt).date()
        except Exception:
            continue
    # Accept sheets like 2025.10.23 or 23.10.2025
    if re.match(r"^\d{4}[./-]\d{1,2}[./-]\d{1,2}$", value):
        digits = re.split(r"[./-]", value)
        try:
            y, m, d = map(int, digits)
            return datetime(year=y, month=m, day=d).date()
        except Exception:
            return None
    if re.match(r"^\d{1,2}[./-]\d{1,2}[./-]\d{4}$", value):
        digits = re.split(r"[./-]", value)
        try:
            d, m, y = map(int, digits)
            return datetime(year=y, month=m, day=d).date()
        except Exception:
            return None
    return None


def detect_schema_for_sheet(ws) -> Dict[str, Any]:
    """
    Heuristic detection:
    - If many label/value pairs in columns A/B → kv mode (single-record form)
    - Else if first non-empty row seems headers → table mode (use first data row)
    Returns schema dict with mode, fields, and positional hints.
    """
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    # Attempt KV mode by scanning first 50 rows, first 4 columns for pairs
    kv_fields = []
    kv_hits = 0
    for r in range(1, min(max_row, 50) + 1):
        label_cell = ws.cell(row=r, column=1)
        value_cell = ws.cell(row=r, column=2)
        label = label_cell.value if label_cell.value is not None else ""
        if isinstance(label, str) and label.strip():
            kv_hits += 1
            kv_fields.append({
                "key": re.sub(r"[^a-zA-Z0-9_]+", "_", label.strip()).strip("_").lower()[:50] or f"field_{r}",
                "label": label.strip(),
                "type": infer_type(value_cell.value),
                "cell": {"row": r, "col": 2},
                "required": False,
            })
    if kv_hits >= 5:
        return {
            "mode": "kv",
            "fields": kv_fields,
        }

    # Header/table mode: find first non-empty row as headers
    header_row_index = None
    headers: List[Tuple[int, str]] = []
    for r in range(1, min(max_row, 30) + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, min(max_col, 50) + 1)]
        non_empty = [v for v in values if v not in (None, "")]
        if len(non_empty) >= 2:
            header_row_index = r
            for c, v in enumerate(values, start=1):
                if v not in (None, ""):
                    label = str(v).strip()
                    headers.append((c, label))
            break
    if header_row_index is not None and headers:
        fields = []
        for c, label in headers:
            key = re.sub(r"[^a-zA-Z0-9_]+", "_", label).strip("_").lower()[:50] or f"col_{c}"
            # Infer type from first data row (header_row_index + 1)
            val = ws.cell(row=min(header_row_index + 1, max_row), column=c).value
            fields.append({
                "key": key,
                "label": label,
                "type": infer_type(val),
                "column": c,
                "required": False,
            })
        return {
            "mode": "table",
            "header_row": header_row_index,
            "data_row": min(header_row_index + 1, max_row),
            "fields": fields,
        }

    # Fallback: basic kv of first 10 rows
    fallback_fields = []
    for r in range(1, min(max_row, 10) + 1):
        label = f"Field {r}"
        fallback_fields.append({
            "key": f"field_{r}",
            "label": label,
            "type": "text",
            "cell": {"row": r, "col": 2},
            "required": False,
        })
    return {"mode": "kv", "fields": fallback_fields}


def infer_type(value: Any) -> str:
    if value is None:
        return "text"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, datetime):
        return "date"
    s = str(value)
    if try_parse_date(s):
        return "date"
    try:
        float(s.replace(",", ""))
        return "number"
    except Exception:
        return "text"


def extract_values(ws, schema: Dict[str, Any]) -> Dict[str, Any]:
    data: Dict[str, Any] = {}
    mode = schema.get("mode", "kv")
    if mode == "kv":
        for f in schema.get("fields", []):
            cell = f.get("cell") or {}
            r = cell.get("row", 1)
            c = cell.get("col", 2)
            data[f["key"]] = ws.cell(row=r, column=c).value
        return data
    # table: take first data row
    data_row = schema.get("data_row", 2)
    for f in schema.get("fields", []):
        c = f.get("column", 1)
        data[f["key"]] = ws.cell(row=data_row, column=c).value
    return data


def build_workbook_schema_and_data(workbook_path: str) -> Dict[str, Dict[str, Any]]:
    """
    Returns mapping: sheet_name -> {schema: {}, values: {}, entry_date: date}
    """
    wb = load_workbook(workbook_path, data_only=True)
    result: Dict[str, Dict[str, Any]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        schema = detect_schema_for_sheet(ws)
        values = extract_values(ws, schema)
        # entry_date from sheet_name or from first date-looking field
        entry_date = try_parse_date(sheet_name)
        if not entry_date:
            for v in values.values():
                if isinstance(v, datetime):
                    entry_date = v.date()
                    break
                if isinstance(v, str):
                    maybe = try_parse_date(v)
                    if maybe:
                        entry_date = maybe
                        break
        result[sheet_name] = {
            "schema": schema,
            "values": values,
            "entry_date": entry_date,
        }
    return result


def write_values_to_workbook(workbook_path: str, sheet_name: str, schema: Dict[str, Any], data: Dict[str, Any]) -> None:
    wb = load_workbook(workbook_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        # Create sheet if missing to avoid data loss
        wb.create_sheet(title=sheet_name)
    ws = wb[sheet_name]
    mode = schema.get("mode", "kv")
    if mode == "kv":
        for f in schema.get("fields", []):
            cell = f.get("cell") or {}
            r = cell.get("row", 1)
            c = cell.get("col", 2)
            ws.cell(row=r, column=c, value=data.get(f["key"]))
    else:
        data_row = schema.get("data_row") or 2
        header_row = schema.get("header_row") or 1
        # Ensure headers exist
        for f in schema.get("fields", []):
            c = f.get("column", 1)
            ws.cell(row=header_row, column=c, value=f.get("label"))
            ws.cell(row=data_row, column=c, value=data.get(f["key"]))
    wb.save(workbook_path)



